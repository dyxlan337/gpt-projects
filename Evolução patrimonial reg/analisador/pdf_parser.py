"""
Parser de extratos PDF da B3 e Banco Inter.
Extrai transações de movimentação de ativos.
"""

import re
from datetime import date, datetime
from typing import List, Optional, Tuple

import pdfplumber

from models import Transacao, TipoMovimentacao, TipoAtivo


# Mapeamento de meses em português
MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3,
    "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Sufixos comuns de tickers de FII
SUFIXOS_FII = ["11", "11B"]
# Sufixos comuns de BDR
SUFIXOS_BDR = ["34", "35", "33", "32"]
# Sufixos ETF
SUFIXOS_ETF = ["11"]
# Tickers conhecidos de ETF
TICKERS_ETF = {
    "BOVA11", "IVVB11", "SMAL11", "HASH11", "XFIX11",
    "BOVV11", "DIVO11", "FIND11", "GOVE11", "MATB11",
    "PIBB11", "SMAC11", "SPXI11", "BRAX11", "ECOO11",
    "ISUS11", "TECK11", "XINA11", "EURP11", "ACWI11",
    "GOLD11", "QBTC11", "BITH11", "ETHE11", "QDFI11",
    "IMAB11", "IRFM11", "B5P211", "IB5M11", "FIXA11",
}

# Tickers conhecidos de FII (parcial, os mais comuns)
TICKERS_FII = {
    "HGLG11", "XPLG11", "XPML11", "KNRI11", "MXRF11",
    "VISC11", "HGBS11", "GGRC11", "HGRE11", "VILG11",
    "BTLG11", "RBRF11", "RBRR11", "KNCR11", "KNIP11",
    "BCFF11", "HFOF11", "VRTA11", "RECR11", "CPTS11",
    "PVBI11", "VGIR11", "TGAR11", "TRXF11", "RZTR11",
    "VGHF11", "HSML11", "BRCR11", "JSRE11", "RBRP11",
}


def parse_numero_br(texto: str) -> float:
    """Converte número no formato brasileiro (1.234,56) para float."""
    if not texto or texto.strip() == "-":
        return 0.0
    texto = texto.strip().replace("R$", "").strip()
    # Remove pontos de milhar e troca vírgula por ponto
    texto = texto.replace(".", "").replace(",", ".")
    # Remove espaços e sinais de menos duplicados
    texto = texto.strip()
    try:
        return float(texto)
    except ValueError:
        return 0.0


def parse_data_br(texto: str) -> Optional[date]:
    """Converte data em formato brasileiro para objeto date.

    Suporta:
    - dd/mm/yyyy
    - dd/mm/yy
    - dd de mês de yyyy
    """
    texto = texto.strip()

    # Formato dd/mm/yyyy ou dd/mm/yy
    match = re.match(r"(\d{2})/(\d{2})/(\d{2,4})", texto)
    if match:
        dia, mes, ano = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if ano < 100:
            ano += 2000
        try:
            return date(ano, mes, dia)
        except ValueError:
            return None

    # Formato "19 de março de 2026"
    match = re.match(
        r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", texto, re.IGNORECASE
    )
    if match:
        dia = int(match.group(1))
        mes_nome = match.group(2).lower()
        ano = int(match.group(3))
        mes = MESES_PT.get(mes_nome)
        if mes:
            try:
                return date(ano, mes, dia)
            except ValueError:
                return None

    return None


def classificar_tipo_ativo(ticker: str, nome_completo: str = "") -> TipoAtivo:
    """Classifica o tipo de ativo baseado no ticker e nome."""
    ticker_upper = ticker.upper().strip()
    nome_upper = nome_completo.upper()

    # ETFs conhecidos
    if ticker_upper in TICKERS_ETF:
        return TipoAtivo.ETF

    # FIIs conhecidos
    if ticker_upper in TICKERS_FII:
        return TipoAtivo.FII

    # Tesouro Direto
    if "TESOURO" in nome_upper or "TESOURO" in ticker_upper:
        return TipoAtivo.TESOURO

    # CDB
    if "CDB" in nome_upper or "CDB" in ticker_upper:
        return TipoAtivo.CDB

    # LCI (word boundary para evitar falsos positivos)
    if re.search(r"\bLCI\b", nome_upper) or ticker_upper.startswith("LCI"):
        return TipoAtivo.LCI

    # LCA (word boundary para evitar "VULCABRAS" etc.)
    if re.search(r"\bLCA\b", nome_upper) or ticker_upper.startswith("LCA"):
        return TipoAtivo.LCA

    # BDR - termina em 34, 35, 33, 32
    base = re.sub(r"\d+[BF]?$", "", ticker_upper)
    sufixo = ticker_upper[len(base):]
    if sufixo in SUFIXOS_BDR and "BDR" in nome_upper:
        return TipoAtivo.BDR
    if "BDR" in nome_upper:
        return TipoAtivo.BDR

    # FII - termina em 11 (e não é ETF)
    if sufixo in SUFIXOS_FII and ("FII" in nome_upper or "FUNDO" in nome_upper):
        return TipoAtivo.FII

    # Ação - padrão para tickers com 4 letras + número
    if re.match(r"^[A-Z]{4}\d{1,2}[BF]?$", ticker_upper):
        if sufixo in ("3", "4", "5", "6"):
            return TipoAtivo.ACAO
        if sufixo in SUFIXOS_FII:
            # Pode ser FII ou ETF, default para FII se tem 11
            return TipoAtivo.FII

    return TipoAtivo.DESCONHECIDO


def classificar_movimentacao(tipo_texto: str) -> TipoMovimentacao:
    """Classifica o tipo de movimentação baseado no texto do extrato."""
    texto = tipo_texto.upper().strip()

    # Cancelado — checar ANTES de dividendo/JCP para pegar "Dividendo - Cancelado"
    if "CANCELAD" in texto:
        return TipoMovimentacao.CANCELADO

    # "COMPRA / VENDA" (renda fixa) — é compra quando o campo Credito/Debito decide;
    # a classificação Credito→COMPRA / Debito→VENDA é feita no parser do XLSX.
    # Aqui retornamos COMPRA como default; o parser override com Credito/Debito.
    if "COMPRA / VENDA" in texto or "COMPRA/VENDA" in texto:
        return TipoMovimentacao.COMPRA

    # Compra / Liquidação de compra
    if any(p in texto for p in [
        "TRANSFERÊNCIA - LIQUIDAÇÃO",
        "TRANSFERENCIA - LIQUIDACAO",
        "COMPRA",
    ]):
        return TipoMovimentacao.COMPRA

    # Venda
    if "VENDA" in texto:
        return TipoMovimentacao.VENDA

    # Dividendos
    if "DIVIDENDO" in texto:
        return TipoMovimentacao.DIVIDENDO

    # JCP - Juros sobre capital próprio
    if "JCP" in texto or "JUROS S/" in texto or "JUROS SOBRE" in texto:
        return TipoMovimentacao.JCP

    # Rendimento
    if "RENDIMENTO" in texto:
        return TipoMovimentacao.RENDIMENTO

    # Bonificação
    if "BONIFICA" in texto:
        return TipoMovimentacao.BONIFICACAO

    # Fração / Leilão de Fração
    is_fracao = "FRAÇÃO" in texto or "FRACAO" in texto or "FRA\x87" in texto
    is_leilao = "LEILÃO" in texto or "LEILAO" in texto or "LEIL" in texto
    if is_fracao or is_leilao:
        if is_leilao:
            return TipoMovimentacao.LEILAO_FRACAO
        return TipoMovimentacao.FRACAO

    # Renda fixa
    if "APLICAÇÃO" in texto or "APLICACAO" in texto:
        return TipoMovimentacao.APLICACAO
    if "RESGATE" in texto:
        return TipoMovimentacao.RESGATE
    if "VENCIMENTO" in texto:
        return TipoMovimentacao.VENCIMENTO

    # Atualização
    if "ATUALIZA" in texto:
        return TipoMovimentacao.ATUALIZACAO

    # Direito de subscrição
    if "SUBSCRI" in texto:
        return TipoMovimentacao.DIREITO_SUBSCRICAO

    # Cessão de direitos
    if "CESSÃO" in texto or "CESSAO" in texto:
        return TipoMovimentacao.CESSAO_DIREITOS

    # Pagamento de juros (renda fixa)
    if "PAGAMENTO" in texto and "JUROS" in texto:
        return TipoMovimentacao.PAGAMENTO_JUROS

    return TipoMovimentacao.IGNORAR


def _extrair_ticker(texto: str) -> str:
    """Extrai o ticker de um texto de descrição do ativo."""
    # Padrão: "PETR4 - PETROBRAS PN" ou "PETR4"
    match = re.match(r"([A-Z]{4}\d{1,2}[BF]?)", texto.upper().strip())
    if match:
        return match.group(1)
    # Tenta encontrar em qualquer posição
    match = re.search(r"\b([A-Z]{4}\d{1,2}[BF]?)\b", texto.upper())
    if match:
        return match.group(1)
    return texto.strip()


def _detectar_direcao_liquidacao(tipo_texto: str) -> TipoMovimentacao:
    """Para 'Transferência - Liquidação', detecta se é compra ou venda
    pelo contexto (campo Debito/Credito ou palavras-chave)."""
    texto = tipo_texto.upper()
    if "DÉBITO" in texto or "DEBITO" in texto:
        return TipoMovimentacao.COMPRA
    if "CRÉDITO" in texto or "CREDITO" in texto:
        return TipoMovimentacao.VENDA
    return TipoMovimentacao.COMPRA  # default


def parse_extrato_b3(caminho_pdf: str) -> List[Transacao]:
    """
    Parse do extrato de movimentação da B3 (CEI).

    O extrato da B3 usa datas em português como cabeçalhos de seção
    ("19 de março de 2026") seguidas de tabelas com colunas:
    Movimentação | Produto | Instituição | Quantidade | Preço Unitário | Valor da Operação

    Detecta pontos coloridos (Credito=verde, Debito=laranja) para
    distinguir compras de vendas em "Transferência - Liquidação".
    """
    transacoes = []

    with pdfplumber.open(caminho_pdf) as pdf:
        texto_completo = ""
        # Mapa de (page_offset_linhas, y_posicao) -> "CREDITO" ou "DEBITO"
        marcadores_direcao = {}
        offset_linhas = 0

        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto:
                # Extrair pontos coloridos desta página
                _extrair_marcadores_pagina(
                    pagina, offset_linhas, marcadores_direcao
                )
                n_linhas = texto.count("\n") + 1
                offset_linhas += n_linhas
                texto_completo += texto + "\n"

        if texto_completo:
            transacoes = _processar_texto_b3(texto_completo, marcadores_direcao)

    return transacoes


# Cores dos pontos no PDF da B3
_COR_CREDITO = (0.0, 0.52941, 0.31373)   # Verde = entrada/compra
_COR_DEBITO = (0.8, 0.26667, 0.0)         # Laranja = saída/venda


def _extrair_marcadores_pagina(
    pagina, offset_linhas: int, marcadores: dict
) -> None:
    """Extrai pontos coloridos (Credito/Debito) e mapeia para linhas do texto.

    Os PDFs da B3 usam pontos coloridos ao lado de cada transação:
    - Verde (0.0, 0.53, 0.31) = Credito = compra/entrada
    - Laranja (0.8, 0.27, 0.0) = Debito = venda/saída
    """
    curves = getattr(pagina, 'curves', [])
    if not curves:
        return

    words = pagina.extract_words()
    if not words:
        return

    # Construir mapa de y -> indice de linha do texto
    # Agrupar words por y para descobrir quais y-positions correspondem
    # a quais linhas do extract_text()
    sorted_words = sorted(words, key=lambda w: (w['top'], w['x0']))
    line_ys = []  # lista de y-positions, uma por linha do texto
    current_y = -100
    for w in sorted_words:
        if abs(w['top'] - current_y) > 3:  # nova linha visual
            line_ys.append(w['top'])
            current_y = w['top']

    for curve in curves:
        color = curve.get('non_stroking_color')
        if not color or not isinstance(color, tuple) or len(color) < 3:
            continue

        r, g = color[0], color[1]
        if abs(r) < 0.05 and abs(g - 0.52941) < 0.05:
            direcao = "CREDITO"
        elif abs(r - 0.8) < 0.05 and abs(g - 0.26667) < 0.05:
            direcao = "DEBITO"
        else:
            continue

        curve_y = curve.get('top', 0)

        # Encontrar a linha visual mais próxima
        melhor_idx = -1
        melhor_dist = float('inf')
        for idx, ly in enumerate(line_ys):
            dist = abs(ly - curve_y)
            if dist < melhor_dist:
                melhor_dist = dist
                melhor_idx = idx

        if melhor_idx >= 0 and melhor_dist < 15:
            linha_global = offset_linhas + melhor_idx
            marcadores[linha_global] = direcao


# Palavras-chave que iniciam uma nova transação
_MOV_KEYWORDS_UPPER = [
    "TRANSFERÊNCIA", "TRANSFERENCIA", "TRANSFER",
    "DIVIDENDO",
    "APLICAÇÃO", "APLICACAO", "APLICA",
    "RESGATE",
    "LEILÃO", "LEILAO", "LEIL",
    "ATUALIZAÇÃO", "ATUALIZACAO", "ATUALIZA",
    "JUROS SOBRE", "JUROS",
    "RENDIMENTO",
    "BONIFICAÇÃO", "BONIFICACAO", "BONIFICA",
    "FRAÇÃO", "FRACAO", "FRA",
    "CESSÃO", "CESSAO", "CESS",
    "DIREITO",
    "PAGAMENTO",
    "VENDA",
    "COMPRA",
]


def _linha_inicia_transacao(linha: str) -> bool:
    """Verifica se a linha inicia uma nova transação."""
    lu = linha.upper()
    for kw in _MOV_KEYWORDS_UPPER:
        if lu.startswith(kw):
            return True
    return False


def _linha_ignorar(linha: str) -> bool:
    """Verifica se a linha é cabeçalho de coluna, header de página ou footer."""
    lu = linha.upper()
    padroes_ignorar = [
        "EXTRATO DE MOVIMENTA",
        "MOVIMENTA\x87\x8BO PRODUTO",      # encoding variante
        "UNITÁRIO", "UNIT\x8BRIO",
        "PREÇO VALOR", "PRE\x87O VALOR",
        "ACESSE INVESTIDOR",
        "FILTROS APLICAD",
        "DATA INICIAL",
        "TIPO DE MOVIMENTA",
        "TIPO DE INVESTIMENTO",
    ]
    for p in padroes_ignorar:
        if p in lu:
            return True
    # Linha que é só "Produto Instituição Quantidade" etc (header de coluna)
    if "PRODUTO" in lu and "QUANTIDADE" in lu:
        return True
    # Linha com CPF
    if "CPF" in lu and "|" in lu:
        return True
    return False


def _buscar_direcao_bloco(
    linhas_nums: list, marcadores: dict
) -> Optional[str]:
    """Busca marcador de direção (CREDITO/DEBITO) em qualquer linha do bloco."""
    for num in linhas_nums:
        direcao = marcadores.get(num)
        if direcao:
            return direcao
    return None


def _processar_texto_b3(
    texto: str, marcadores_direcao: dict = None
) -> List[Transacao]:
    """Processa texto do PDF da B3 onde datas são cabeçalhos de seção em português.

    Args:
        texto: texto completo de todas as páginas concatenado
        marcadores_direcao: mapa de linha_global -> "CREDITO"/"DEBITO"
    """
    transacoes = []
    linhas = texto.split("\n")

    if marcadores_direcao is None:
        marcadores_direcao = {}

    # Padrão de data em português: "19 de março de 2026"
    padrao_data_pt = re.compile(
        r"^(\d{1,2})\s+de\s+\w+\s+de\s+(\d{4})$", re.IGNORECASE
    )

    data_atual = None
    # lista de (date, [linhas_do_bloco], direcao_ou_None)
    blocos = []
    bloco_atual = []
    linhas_nums_bloco = []  # números de linha do bloco atual

    for num_linha, linha in enumerate(linhas):
        s = linha.strip()
        if not s:
            continue

        # Data de seção?
        if padrao_data_pt.match(s):
            # Salva bloco anterior
            if bloco_atual and data_atual:
                direcao = _buscar_direcao_bloco(linhas_nums_bloco, marcadores_direcao)
                blocos.append((data_atual, bloco_atual, direcao))
                bloco_atual = []
                linhas_nums_bloco = []
            data_atual = parse_data_br(s)
            continue

        # Linha de cabeçalho/rodapé? Pula
        if _linha_ignorar(s):
            continue

        # Linha inicia nova transação?
        if _linha_inicia_transacao(s):
            if bloco_atual and data_atual:
                direcao = _buscar_direcao_bloco(linhas_nums_bloco, marcadores_direcao)
                blocos.append((data_atual, bloco_atual, direcao))
                bloco_atual = []
                linhas_nums_bloco = []

        # Só coleta linhas se já temos uma data
        if data_atual:
            bloco_atual.append(s)
            linhas_nums_bloco.append(num_linha)

    # Último bloco
    if bloco_atual and data_atual:
        direcao = _buscar_direcao_bloco(linhas_nums_bloco, marcadores_direcao)
        blocos.append((data_atual, bloco_atual, direcao))

    # Parseia cada bloco em transação
    for dt, linhas_bloco, direcao in blocos:
        trans = _parse_bloco_b3(dt, linhas_bloco, direcao)
        if trans:
            transacoes.append(trans)

    return transacoes


def _parse_bloco_b3(
    data: date, linhas: list, direcao: str = None
) -> Optional[Transacao]:
    """Extrai uma Transacao de um bloco de 2-3 linhas de texto.

    Args:
        direcao: "CREDITO" (compra) ou "DEBITO" (venda), detectado pelos
                 pontos coloridos do PDF. None se não disponível.
    """
    texto = " ".join(linhas)
    texto_upper = texto.upper()

    # 1) Tipo de movimentação
    tipo = classificar_movimentacao(texto)

    # Para "Transferência - Liquidação", usar direção do ponto colorido
    is_liquidacao = (
        "LIQUIDAÇÃO" in texto_upper or "LIQUIDACAO" in texto_upper
        or "LIQUIDA" in texto_upper
    )
    if is_liquidacao:
        if direcao == "DEBITO":
            tipo = TipoMovimentacao.VENDA
        elif direcao == "CREDITO":
            tipo = TipoMovimentacao.COMPRA
        else:
            # Sem marcador colorido, mantém heurística anterior
            tipo = _detectar_direcao_liquidacao(texto)

    # 2) Extrair ticker
    ticker = ""
    nome_completo = ""

    # Padrão de ação/FII/BDR/ETF: XXXX3, XPML11, NUBR33, etc.
    ticker_match = re.search(r'\b([A-Z]{4}\d{1,2}[BF]?)\b', texto_upper)
    if ticker_match:
        ticker = ticker_match.group(1)
        # Extrair nome completo: "TICKER - NOME..."
        nome_match = re.search(
            rf'{re.escape(ticker)}\s*-\s*([A-ZÁÉÍÓÚÃÕÂÊÔÇ][A-ZÁÉÍÓÚÃÕÂÊÔÇ\s/.,]+?)(?=\s+(?:INTER|BANCO|XP|CLEAR|NU|RICO)\s)',
            texto_upper
        )
        if nome_match:
            nome_completo = f"{ticker} - {nome_match.group(1).strip()}"
        else:
            nome_completo = ticker

    # Padrão CDB/LCI/LCA: "CDB - CDB326844PK" ou "LCI - ..."
    if not ticker:
        renda_fixa_match = re.search(
            r'((?:CDB|LCI|LCA)\s*-\s*\S+)',
            texto, re.IGNORECASE
        )
        if renda_fixa_match:
            ticker = re.sub(r'\s+', '', renda_fixa_match.group(1)).upper()
            nome_completo = ticker

    if not ticker:
        return None

    # 3) Extrair valores numéricos (quantidade, preço, valor)
    quantidade, preco, valor = _extrair_numeros_bloco(texto, ticker)

    # Ignorar transações sem valores relevantes (tipo Atualização com tudo zero)
    if quantidade == 0 and valor == 0:
        return None

    # Se não tem valor mas tem qtd e preço, calcula
    if valor == 0 and quantidade > 0 and preco > 0:
        valor = quantidade * preco

    # Se não tem preço mas tem qtd e valor, calcula
    if preco == 0 and quantidade > 0 and valor > 0:
        preco = valor / quantidade

    tipo_ativo = classificar_tipo_ativo(ticker, nome_completo)

    return Transacao(
        data=data,
        tipo=tipo,
        ticker=ticker,
        nome_completo=nome_completo,
        instituicao="",
        quantidade=abs(quantidade),
        preco_unitario=abs(preco),
        valor_operacao=abs(valor),
        tipo_ativo=tipo_ativo,
    )


def _extrair_numeros_bloco(texto: str, ticker: str) -> Tuple[float, float, float]:
    """Extrai (quantidade, preço_unitário, valor_operação) do texto de um bloco.

    Estratégia:
    1. Remove tickers, códigos CDB e marcadores R$ do texto
    2. Extrai todos os números na ordem em que aparecem
    3. Usa a relação qty * preço ≈ valor para identificar corretamente
    """
    # Contar "R$ -" (valores zero)
    zeros_rs = len(re.findall(r'R\$\s*-', texto))

    # Limpar texto: remover tickers e códigos de renda fixa
    texto_limpo = re.sub(r'\b[A-Z]{3,6}\d{1,2}[BF]?\b', ' ', texto)
    texto_limpo = re.sub(r'CDB[-\s]?\w+', ' ', texto_limpo, flags=re.IGNORECASE)
    texto_limpo = re.sub(r'LCI[-\s]?\w+', ' ', texto_limpo, flags=re.IGNORECASE)
    texto_limpo = re.sub(r'LCA[-\s]?\w+', ' ', texto_limpo, flags=re.IGNORECASE)

    # Remover marcadores R$ (manter os números)
    texto_limpo = re.sub(r'R\$', ' ', texto_limpo)

    # Extrair todos os números na ordem de aparição
    numeros = []
    for m in re.finditer(r'(\d[\d.,]*\d|\d)', texto_limpo):
        val = parse_numero_br(m.group())
        numeros.append(val)

    # Filtrar zeros
    numeros_pos = [n for n in numeros if n > 0]

    if not numeros_pos:
        return (0.0, 0.0, 0.0)

    if len(numeros_pos) == 1:
        return (numeros_pos[0], 0.0, 0.0)

    if len(numeros_pos) == 2:
        a, b = numeros_pos
        if zeros_rs >= 1:
            return (a, 0.0, b)
        # Sem R$-, assume qty e valor
        return (a, b / a if a > 0 else 0.0, b)

    # 3+ números: encontrar trinca (q, p, v) onde q * p ≈ v
    # No PDF da B3, a ordem no texto pode ser (qty, preço, valor) OU (qty, valor, preço)
    # porque as colunas preço e valor podem cair em linhas diferentes
    tolerancia = lambda esperado, real: abs(esperado - real) < 0.05 + abs(real) * 0.02

    # Tenta todas as combinações de 3 números (preservando ordem)
    for i in range(len(numeros_pos)):
        for j in range(i + 1, len(numeros_pos)):
            for k in range(j + 1, len(numeros_pos)):
                a, b, c = numeros_pos[i], numeros_pos[j], numeros_pos[k]
                # Caso 1: (qty, preço, valor) → a*b ≈ c
                if c > 0 and tolerancia(a * b, c):
                    return (a, b, c)
                # Caso 2: (qty, valor, preço) → a*c ≈ b
                if b > 0 and tolerancia(a * c, b):
                    return (a, c, b)

    # Se não encontrou trinca perfeita, tenta os últimos 3
    a, b, c = numeros_pos[-3], numeros_pos[-2], numeros_pos[-1]
    if c > 0 and tolerancia(a * b, c):
        return (a, b, c)
    if b > 0 and tolerancia(a * c, b):
        return (a, c, b)

    # Fallback: primeiro número = qty, últimos dois ordenados como (preço, valor)
    p, v = sorted([numeros_pos[-2], numeros_pos[-1]])
    return (numeros_pos[0], p, v)


def parse_csv_manual(caminho_csv: str, separador: str = ";") -> List[Transacao]:
    """
    Parse de planilha CSV manual com formato:
    Data;Ticker;Tipo;Quantidade;Preço;Instituição

    ou formato simplificado:
    Data;Ticker;Quantidade;Preço
    """
    import csv

    transacoes = []

    with open(caminho_csv, "r", encoding="utf-8-sig") as f:
        leitor = csv.reader(f, delimiter=separador)
        header = next(leitor, None)
        if not header:
            return []

        # Normaliza header
        header_norm = [h.upper().strip() for h in header]

        # Mapeia colunas
        col = {}
        for i, h in enumerate(header_norm):
            if "DATA" in h:
                col["data"] = i
            elif "TICKER" in h or "ATIVO" in h or "CÓDIGO" in h or "CODIGO" in h:
                col["ticker"] = i
            elif "TIPO" in h or "MOVIMENTA" in h:
                col["tipo"] = i
            elif "QTD" in h or "QUANTIDADE" in h:
                col["quantidade"] = i
            elif "PREÇO" in h or "PRECO" in h or "VALOR UNIT" in h:
                col["preco"] = i
            elif "INSTITUI" in h or "CORRETORA" in h:
                col["instituicao"] = i
            elif "VALOR" in h and "valor" not in col:
                col["valor"] = i
            elif "NOME" in h:
                col["nome"] = i

        if "data" not in col or "ticker" not in col:
            raise ValueError(
                "CSV deve ter pelo menos colunas 'Data' e 'Ticker'. "
                f"Colunas encontradas: {header}"
            )

        for linha in leitor:
            if not linha or all(c.strip() == "" for c in linha):
                continue

            try:
                data = parse_data_br(linha[col["data"]])
                if not data:
                    continue

                ticker = linha[col["ticker"]].strip().upper()
                if not ticker:
                    continue

                tipo_texto = linha[col["tipo"]].strip() if "tipo" in col else "COMPRA"
                tipo = classificar_movimentacao(tipo_texto)

                qtd = parse_numero_br(linha[col["quantidade"]]) if "quantidade" in col else 0.0
                preco = parse_numero_br(linha[col["preco"]]) if "preco" in col else 0.0
                valor = parse_numero_br(linha[col.get("valor", -1)]) if "valor" in col else qtd * preco
                inst = linha[col["instituicao"]].strip() if "instituicao" in col else ""
                nome = linha[col["nome"]].strip() if "nome" in col else ticker

                tipo_ativo = classificar_tipo_ativo(ticker, nome)

                transacoes.append(Transacao(
                    data=data,
                    tipo=tipo,
                    ticker=ticker,
                    nome_completo=nome,
                    instituicao=inst,
                    quantidade=abs(qtd),
                    preco_unitario=abs(preco),
                    valor_operacao=abs(valor),
                    tipo_ativo=tipo_ativo,
                ))
            except (IndexError, ValueError):
                continue

    return transacoes


def parse_xlsx_b3(caminho_xlsx: str) -> List[Transacao]:
    """
    Parse da planilha XLSX exportada pelo site da B3 (CEI).

    Colunas esperadas:
    Entrada/Saída | Data | Movimentação | Produto | Instituição |
    Quantidade | Preço unitário | Valor da Operação
    """
    from datetime import datetime as dt_class
    from openpyxl import load_workbook

    transacoes = []
    wb = load_workbook(caminho_xlsx, data_only=True)
    ws = wb.active

    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_data) < 2:
        return []

    # Mapear colunas pelo header
    header = [str(h).upper().strip() if h else "" for h in rows_data[0]]
    col = {}
    for i, h in enumerate(header):
        if "ENTRADA" in h or "SAÍDA" in h or "SAIDA" in h:
            col["direcao"] = i
        elif "DATA" in h:
            col["data"] = i
        elif "MOVIMENTA" in h:
            col["movimentacao"] = i
        elif "PRODUTO" in h:
            col["produto"] = i
        elif "INSTITUI" in h:
            col["instituicao"] = i
        elif "QUANTIDADE" in h or "QTD" in h:
            col["quantidade"] = i
        elif "PREÇO" in h or "PRECO" in h or "UNITÁRIO" in h or "UNITARIO" in h:
            col["preco"] = i
        elif "VALOR" in h and ("OPERA" in h or "TOTAL" in h):
            col["valor"] = i
        elif "VALOR" in h and "valor" not in col:
            col["valor"] = i

    if "data" not in col:
        raise ValueError(
            f"XLSX deve ter coluna 'Data'. Colunas encontradas: {header}"
        )

    for row in rows_data[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            # Data
            data_raw = row[col["data"]]
            if isinstance(data_raw, dt_class):
                data_val = data_raw.date()
            elif isinstance(data_raw, date):
                data_val = data_raw
            else:
                data_val = parse_data_br(str(data_raw))
            if not data_val:
                continue

            # Direção (Credito/Debito)
            direcao = ""
            if "direcao" in col and row[col["direcao"]]:
                direcao = str(row[col["direcao"]]).upper().strip()

            # Movimentação
            mov_texto = ""
            if "movimentacao" in col and row[col["movimentacao"]]:
                mov_texto = str(row[col["movimentacao"]]).strip()
            tipo = classificar_movimentacao(mov_texto)

            # Para movimentações ambíguas, usar Credito/Debito para decidir
            mov_upper = mov_texto.upper()
            usa_direcao = (
                "LIQUIDAÇÃO" in mov_upper or "LIQUIDACAO" in mov_upper
                or "LIQUIDA" in mov_upper
                or "COMPRA / VENDA" in mov_upper or "COMPRA/VENDA" in mov_upper
            )
            if usa_direcao:
                if "DEBITO" in direcao or "DÉBITO" in direcao:
                    tipo = TipoMovimentacao.VENDA
                elif "CREDITO" in direcao or "CRÉDITO" in direcao:
                    tipo = TipoMovimentacao.COMPRA

            # Produto / Ticker
            produto_texto = ""
            if "produto" in col and row[col["produto"]]:
                produto_texto = str(row[col["produto"]]).strip()
            ticker = _extrair_ticker(produto_texto)
            nome_completo = produto_texto

            if not ticker:
                continue

            # Instituição
            instituicao = ""
            if "instituicao" in col and row[col["instituicao"]]:
                instituicao = str(row[col["instituicao"]]).strip()

            # Quantidade
            quantidade = 0.0
            if "quantidade" in col and row[col["quantidade"]] is not None:
                raw = row[col["quantidade"]]
                quantidade = float(raw) if isinstance(raw, (int, float)) else parse_numero_br(str(raw))

            # Preço unitário
            preco = 0.0
            if "preco" in col and row[col["preco"]] is not None:
                raw = row[col["preco"]]
                if isinstance(raw, (int, float)):
                    preco = float(raw)
                else:
                    preco = parse_numero_br(str(raw).replace("R$", "").strip())

            # Valor operação
            valor = 0.0
            if "valor" in col and row[col["valor"]] is not None:
                raw = row[col["valor"]]
                if isinstance(raw, (int, float)):
                    valor = float(raw)
                else:
                    valor = parse_numero_br(str(raw).replace("R$", "").strip())

            if valor == 0 and quantidade > 0 and preco > 0:
                valor = quantidade * preco
            if preco == 0 and quantidade > 0 and valor > 0:
                preco = valor / quantidade

            tipo_ativo = classificar_tipo_ativo(ticker, nome_completo)

            transacoes.append(Transacao(
                data=data_val,
                tipo=tipo,
                ticker=ticker,
                nome_completo=nome_completo,
                instituicao=instituicao,
                quantidade=abs(quantidade),
                preco_unitario=abs(preco),
                valor_operacao=abs(valor),
                tipo_ativo=tipo_ativo,
            ))
        except (IndexError, ValueError, TypeError):
            continue

    return transacoes


# ---------------------------------------------------------------------------
# Parser do relatório PDF do Banco Inter
# ---------------------------------------------------------------------------

def _parse_data_inter(texto: str) -> Optional[date]:
    """Parse de datas em português com encoding possivelmente corrompido."""
    texto = texto.strip()
    m = re.match(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", texto, re.IGNORECASE)
    if not m:
        return None
    dia, mes_nome, ano = int(m.group(1)), m.group(2).lower(), int(m.group(3))
    mes = MESES_PT.get(mes_nome)
    if not mes:
        for k, v in MESES_PT.items():
            if mes_nome in k or k in mes_nome:
                mes = v
                break
    if mes:
        try:
            return date(ano, mes, dia)
        except ValueError:
            return None
    return None


def _is_inter_boundary(line, date_pat, header_pat, page_pat, rf_pat, rv_pat, td_pat):
    """Verifica se a linha é uma fronteira entre seções."""
    if not line:
        return True
    if date_pat.match(line) or header_pat.search(line) or page_pat.match(line):
        return True
    if rf_pat.match(line) or rv_pat.match(line) or td_pat.match(line):
        return True
    if line.strip() == "Tesouro":
        return True
    if re.match(r"^(LC[IA]|CDB)\s", line, re.IGNORECASE):
        return True
    return False


def parse_inter_pdf(caminho_pdf: str) -> List[Transacao]:
    """
    Parse do relatório de movimentações do Banco Inter (PDF).

    Formato: datas em português como cabeçalhos de seção, seguidas de linhas
    com: Produto | Ativo | Código | Quantidade | Valor | IOF | IR | Valor Líquido

    Convenção do PDF Inter:
    - \\x00 antes do número de quantidade = SAÍDA (resgate/venda)
    - Sem \\x00 antes da quantidade = ENTRADA (aplicação/compra)
    """
    transacoes = []

    date_pat = re.compile(r"^(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})$", re.IGNORECASE)
    header_pat = re.compile(r"Produto\s+Ativo|C.digo\s*Ativo", re.IGNORECASE)
    page_pat = re.compile(r"CPF\s+\d|Extrato de movimenta", re.IGNORECASE)
    rf_pat = re.compile(r"Renda\s+Fixa\s+(.*)", re.IGNORECASE)
    rv_pat = re.compile(r"Renda\s+Vari.vel\s*(.*)", re.IGNORECASE)
    td_pat = re.compile(r"Tesouro\s+Direto\s+(.*)", re.IGNORECASE)

    def is_boundary(ln):
        return _is_inter_boundary(ln, date_pat, header_pat, page_pat, rf_pat, rv_pat, td_pat)

    all_lines = []
    with pdfplumber.open(caminho_pdf) as pdf:
        for pg_num in range(3, len(pdf.pages)):
            page = pdf.pages[pg_num]
            text = page.extract_text()
            if text:
                for line in text.split("\n"):
                    all_lines.append(line)

    current_date = None
    pending_prefix = None
    i = 0

    while i < len(all_lines):
        line = all_lines[i].strip()

        if not line or page_pat.match(line) or header_pat.search(line):
            i += 1
            continue

        dm = date_pat.match(line)
        if dm:
            d = _parse_data_inter(dm.group(1))
            if d:
                current_date = d
            pending_prefix = None
            i += 1
            continue

        if not current_date:
            i += 1
            continue

        # Prefixo multi-linha (ex: "LCA IPCA 252", "CDB PRE 252")
        if re.match(r"^(LC[IA]|CDB)\s", line, re.IGNORECASE) and not rf_pat.match(line):
            pending_prefix = line
            i += 1
            continue

        if line.strip() == "Tesouro":
            pending_prefix = "Tesouro"
            i += 1
            continue

        # === RENDA FIXA ===
        rf_m = rf_pat.match(line)
        if rf_m:
            rest = rf_m.group(1)
            while i + 1 < len(all_lines):
                next_line = all_lines[i + 1].strip()
                if is_boundary(next_line):
                    break
                rest += " " + next_line
                i += 1

            full_product = (pending_prefix + " " if pending_prefix else "") + rest
            pending_prefix = None

            qty_m = re.search(r"(\x00)?(\d[\d]*\.?\d*)\s+\x00?R\$\s+([\d.,]+)", rest)
            qty_real = 1.0
            if not qty_m:
                dash_m = re.search(r"-\s+\x00?R\$\s+([\d.,]+)", rest)
                if dash_m:
                    valor = parse_numero_br(dash_m.group(1))
                    is_saida = False
                else:
                    i += 1
                    continue
            else:
                is_saida = bool(qty_m.group(1))
                qty_str = qty_m.group(2)
                valor = parse_numero_br(qty_m.group(3))
                qty_real = float(qty_str)
                if qty_str in ("3",) and not is_saida:
                    rest2 = rest[qty_m.end():]
                    qty_m2 = re.search(r"(\x00)?(\d[\d]*\.?\d*)\s+\x00?R\$\s+([\d.,]+)", rest2)
                    if qty_m2:
                        is_saida = bool(qty_m2.group(1))
                        qty_str = qty_m2.group(2)
                        valor = parse_numero_br(qty_m2.group(3))
                        qty_real = float(qty_str)

            code_m = re.search(r"(CDB\w+|LC[IA]\w+|\d{2}[A-Z]\d{7,})", full_product)
            code = code_m.group(1) if code_m else "RF-DESCONHECIDO"

            fp_upper = full_product.upper()
            if "LCI" in fp_upper:
                nome = f"LCI-{code}"
                tipo_ativo = TipoAtivo.LCI
            elif "LCA" in fp_upper:
                nome = f"LCA-{code}"
                tipo_ativo = TipoAtivo.LCA
            else:
                nome = f"CDB-{code}"
                tipo_ativo = TipoAtivo.CDB

            tipo_mov = TipoMovimentacao.RESGATE if is_saida else TipoMovimentacao.APLICACAO

            # Usar quantidade real do PDF para FIFO funcionar em resgates parciais
            preco_unit = valor / qty_real if qty_real > 0 else valor

            transacoes.append(Transacao(
                data=current_date,
                tipo=tipo_mov,
                ticker=nome,
                nome_completo=full_product.strip()[:100],
                instituicao="INTER",
                quantidade=qty_real,
                preco_unitario=preco_unit,
                valor_operacao=valor,
                tipo_ativo=tipo_ativo,
            ))
            i += 1
            continue

        # === TESOURO DIRETO ===
        td_m = td_pat.match(line)
        if td_m:
            rest = td_m.group(1)
            pending_prefix = None
            while i + 1 < len(all_lines):
                next_line = all_lines[i + 1].strip()
                if is_boundary(next_line):
                    break
                rest += " " + next_line
                i += 1

            qty_m = re.search(r"(\x00)?(\d+\.\d+)\s+\x00?R\$\s+([\d.,]+)", rest)
            if qty_m:
                is_saida = bool(qty_m.group(1))
                qty = float(qty_m.group(2))
                valor = parse_numero_br(qty_m.group(3))

                nome_td = "Tesouro Direto"
                td_nome_m = re.search(r"(Prefixado|Selic|IPCA)[^\d]*(\d{4})?", rest, re.IGNORECASE)
                if td_nome_m:
                    nome_td = f"Tesouro {td_nome_m.group(1)}"
                    if td_nome_m.group(2):
                        nome_td += f" {td_nome_m.group(2)}"

                tipo_mov = TipoMovimentacao.VENDA if is_saida else TipoMovimentacao.COMPRA

                transacoes.append(Transacao(
                    data=current_date,
                    tipo=tipo_mov,
                    ticker=nome_td.upper().replace(" ", "_"),
                    nome_completo=nome_td,
                    instituicao="INTER",
                    quantidade=qty,
                    preco_unitario=valor / qty if qty > 0 else valor,
                    valor_operacao=valor,
                    tipo_ativo=TipoAtivo.TESOURO,
                ))
            i += 1
            continue

        # === RENDA VARIÁVEL ===
        rv_m = rv_pat.match(line)
        if rv_m:
            rest = rv_m.group(1)
            pending_prefix = None
            while i + 1 < len(all_lines):
                next_line = all_lines[i + 1].strip()
                if is_boundary(next_line):
                    break
                rest += " " + next_line
                i += 1

            if "US$" in rest or not rest.strip():
                i += 1
                continue

            ticker_m = re.search(
                r"([A-Z]{4}\d{1,2}[F]?)\s+-\s+(\x00)?(\d+\.\d+)\s+\x00?R\$\s+([\d.,]+)",
                rest,
            )
            if ticker_m:
                ticker = ticker_m.group(1)
                is_saida = bool(ticker_m.group(2))
                qty = float(ticker_m.group(3))
                valor = parse_numero_br(ticker_m.group(4))

                ticker_limpo = ticker.rstrip("F") if ticker.endswith("F") and len(ticker) > 5 else ticker
                tipo_mov = TipoMovimentacao.VENDA if is_saida else TipoMovimentacao.COMPRA
                tipo_ativo = classificar_tipo_ativo(ticker_limpo)

                transacoes.append(Transacao(
                    data=current_date,
                    tipo=tipo_mov,
                    ticker=ticker_limpo,
                    nome_completo=ticker,
                    instituicao="INTER",
                    quantidade=qty,
                    preco_unitario=valor / qty if qty > 0 else valor,
                    valor_operacao=valor,
                    tipo_ativo=tipo_ativo,
                ))
            i += 1
            continue

        i += 1

    return transacoes
